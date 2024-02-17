# -*- coding:UTF-8 -*-

import os
import numpy as np

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font

# author:Zhiheng Feng
# contact: fzhsjtu@foxmail.com
# datetime:2021/10/21 19:46
# software: PyCharm

"""
文件说明：
    读取eval文件中的输出文件
"""


class SaveExcel(object):
    def __init__(self, test_list, root_path, excel_name='output'):
        self.test_list = test_list
        self.save_path = root_path
        excel_path = os.path.join(root_path, '{}.xlsx'.format(excel_name))
        self.excel_path = excel_path
        self.creat_excel()

    def creat_excel(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'sheet1'
        font = Font(bold=True)
        cell0 = sheet.cell(row=1, column=1)
        cell0.value = 'epoch'
        cell0.font = font
        alignment = Alignment(horizontal="center", vertical="center")
        for i, item in enumerate(self.test_list, 1):
            cell1 = sheet.cell(row=1, column=i * 2)
            cell2 = sheet.cell(row=1, column=i * 2 + 1)
            cell1.value = '{:02d} RMSE'.format(item)
            cell2.value = '{:02d} error'.format(item)
            cell1.font = font
            cell2.font = font
            cell1.alignment = alignment
            cell2.alignment = alignment
        workbook.save(filename=self.excel_path)

    def update(self, eval_dir: str, read_file_name: str = 'output'):
        col_len = len(self.test_list)
        workbook = load_workbook(filename=self.excel_path)
        sheet = workbook.active
        sheet.column_dimensions.width = 30
        sheet.column_dimensions['A'].width = 9
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        mean_list = []
        mean_error_list = []
        for col, item in enumerate(self.test_list, 1):
            txt_path = os.path.join(eval_dir, 'translonet_{:02d}'.format(item), '{}.txt'.format(read_file_name))
            if not os.path.exists(txt_path):
                continue
            epoch_list = []
            RMSE_list = []
            error_list = []

            with open(txt_path, 'r') as f:
                data = f.readlines()
                for row in range(int(len(data) / 3)):
                    index1 = data[row * 3].index(':') + 2
                    index2 = data[row * 3 + 1].index(':') + 2
                    index3 = data[row * 3 + 2].index(':') + 2
                    ep = data[row * 3][index1:].strip()
                    rmse = data[row * 3 + 1][index2:].strip()
                    error = data[row * 3 + 2][index3:].strip()
                    cell0 = sheet.cell(row=row + 2, column=1)
                    cell1 = sheet.cell(row=row + 2, column=col * 2)
                    cell2 = sheet.cell(row=row + 2, column=col * 2 + 1)
                    cell0.value = int(ep)
                    cell0.alignment = alignment
                    cell1.value = float(rmse)
                    cell1.alignment = alignment
                    cell2.value = float(error) * 100
                    cell2.alignment = alignment
                    epoch_list.append(int(ep))
                    RMSE_list.append(float(rmse))
                    error_list.append(float(error) * 100)

            mean_list.append(RMSE_list)
            mean_error_list.append(error_list)
            min_RMSE = min(RMSE_list)
            min_index = RMSE_list.index(min_RMSE)

            cell = sheet.cell(row=1, column=col_len * 2 + 2 + col)
            cell.value = '{:02d}'.format(item)
            cell.alignment = alignment

            cell_min = sheet.cell(row=2, column=col_len * 2 + 2 + col)
            cell_min.value = '{:d}: {:.4f}'.format(epoch_list[min_index], min_RMSE)
            cell_min.alignment = alignment

            min_error = min(error_list)
            min_index = error_list.index(min_error)
            cell_min = sheet.cell(row=3, column=col_len * 2 + 2 + col)
            cell_min.value = '{:d}: {:.4f}'.format(epoch_list[min_index], min_error)
            cell_min.alignment = alignment

        mean_array = np.array(mean_list)
        mean = np.mean(mean_array, axis=0)
        min_mean = min(mean)
        min_index = np.where(mean == min_mean)[-1][-1]
        cell = sheet.cell(row=1, column=col_len * 3 + 3)
        cell.value = 'mean_min'
        cell.alignment = alignment
        cell_min = sheet.cell(row=2, column=col_len * 3 + 3)
        cell_min.value = '{:d}: {:.4f}'.format(epoch_list[min_index], min_mean)
        cell_min.alignment = alignment

        mean_array = np.array(mean_error_list)
        mean = np.mean(mean_array, axis=0)
        min_mean = min(mean)
        min_index = np.where(mean == min_mean)[-1][-1]
        cell_min = sheet.cell(row=3, column=col_len * 3 + 3)
        cell_min.value = '{:d}: {:.4f}'.format(epoch_list[min_index], min_mean)
        cell_min.alignment = alignment

        workbook.save(filename=self.excel_path)
