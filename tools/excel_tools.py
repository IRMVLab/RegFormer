# -*- coding:UTF-8 -*-

import os
import numpy as np

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font

# author:Zhiheng Feng
# contact: fzhsjtu@foxmail.com
# datetime:2021/10/21 19:46
# software: PyCharm

"""
文件说明：
    读取eval文件中的输出文件
"""


class SaveExcel(object):
    def __init__(self, test_list, root_path, excel_name='reg_output'):
        self.test_list = test_list
        self.save_path = root_path
        excel_path = os.path.join(root_path, '{}.xlsx'.format(excel_name))
        self.excel_path = excel_path
        self.creat_excel()

    def creat_excel(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'sheet1'
        font = Font(bold=True)
        cell0 = sheet.cell(row=1, column=1)
        cell0.value = 'epoch'
        cell0.font = font
        alignment = Alignment(horizontal="center", vertical="center")

        cell1 = sheet.cell(row=1, column= 2)
        cell2 = sheet.cell(row=1, column= 3)
        cell3 = sheet.cell(row=1, column= 4)
        cell4 = sheet.cell(row=1, column= 5)
        cell5 = sheet.cell(row=1, column= 6)

        cell1.value = 'RR'
        cell2.value = 'TM'
        cell3.value = 'TS'
        cell4.value = 'RM'
        cell5.value = 'RS'
        cell1.font = font
        cell2.font = font
        cell3.font = font
        cell4.font = font
        cell5.font = font
        cell1.alignment = alignment
        cell2.alignment = alignment
        cell3.alignment = alignment
        cell4.alignment = alignment
        cell5.alignment = alignment
        workbook.save(filename=self.excel_path)

    def update(self, eval_dir: str, read_file_name: str = 'reg_output'):

        workbook = load_workbook(filename=self.excel_path)
        sheet = workbook.active
        sheet.column_dimensions.width = 30
        sheet.column_dimensions['A'].width = 9
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        mean_RR = []
        mean_TM = []
        mean_TS = []
        mean_RM = []
        mean_RS = []

        txt_path = os.path.join(eval_dir, 'translonet_nuscenes', '{}.txt'.format(read_file_name))
        # if not os.path.exists(txt_path): continue
        epoch_list = []
        RR_list = []
        TM_list = []
        TS_list = []
        RM_list = []
        RS_list = []

        with open(txt_path, 'r') as f:
                data = f.readlines()
                for row in range(int(len(data) / 6)):
                    index1 = data[row * 6].index(':') + 2
                    index2 = data[row * 6 + 1].index(':') + 2
                    index3 = data[row * 6 + 2].index(':') + 2
                    index4 = data[row * 6 + 3].index(':') + 2
                    index5 = data[row * 6 + 4].index(':') + 2
                    index6 = data[row * 6 + 5].index(':') + 2
                    ep = data[row * 6][index1:].strip()
                    RR = data[row * 6 + 1][index2:].strip()
                    TM = data[row * 6 + 2][index3:].strip()
                    TS = data[row * 6 + 3][index4:].strip()
                    RM = data[row * 6 + 4][index5:].strip()
                    RS = data[row * 6 + 5][index6:].strip()

                    cell0 = sheet.cell(row=row + 2, column=1)
                    cell1 = sheet.cell(row=row + 2, column=2)
                    cell2 = sheet.cell(row=row + 2, column=3)
                    cell3 = sheet.cell(row=row + 2, column=4)
                    cell4 = sheet.cell(row=row + 2, column=5)
                    cell5 = sheet.cell(row=row + 2, column=6)

                    cell0.value = int(ep)
                    cell0.alignment = alignment
                    cell1.value = float(RR)
                    cell1.alignment = alignment
                    cell2.value = float(TM)
                    cell2.alignment = alignment
                    cell3.value = float(TS)
                    cell3.alignment = alignment
                    cell4.value = float(RM)
                    cell4.alignment = alignment
                    cell5.value = float(RS)
                    cell5.alignment = alignment

                    epoch_list.append(int(ep))
                    RR_list.append(float(RR))
                    TM_list.append(float(TM))
                    TS_list.append(float(TS))
                    RM_list.append(float(RM))
                    RS_list.append(float(RS))


        cell = sheet.cell(row=1, column=7 )
        cell.value = 'min'
        cell.alignment = alignment

        min_RR = max(RR_list)
        min_index = RR_list.index(min_RR)
        cell_min = sheet.cell(row=2, column=7)
        cell_min.value = '{:d}: {:.4f}'.format(epoch_list[min_index], min_RR)
        cell_min.alignment = alignment

        min_TM = min(TM_list)
        min_index = TM_list.index(min_TM)
        cell_min = sheet.cell(row=3, column=7)
        cell_min.value = '{:d}: {:.4f}'.format(epoch_list[min_index], min_TM)
        cell_min.alignment = alignment

        min_TS = min(TS_list)
        min_index = TS_list.index(min_TS)
        cell_min = sheet.cell(row=4, column=7)
        cell_min.value = '{:d}: {:.4f}'.format(epoch_list[min_index], min_TS)
        cell_min.alignment = alignment

        min_RM = min(RM_list)
        min_index = RM_list.index(min_RM)
        cell_min = sheet.cell(row=5, column=7)
        cell_min.value = '{:d}: {:.4f}'.format(epoch_list[min_index], min_RM)
        cell_min.alignment = alignment

        min_RS = min(RS_list)
        min_index = RS_list.index(min_RS)
        cell_min = sheet.cell(row=6, column=7)
        cell_min.value = '{:d}: {:.4f}'.format(epoch_list[min_index], min_RS)
        cell_min.alignment = alignment



        workbook.save(filename=self.excel_path)
